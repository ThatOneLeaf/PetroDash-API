import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import os

def generate_power_report_template(logo_path, company, powerplant, metric, output_file="Power_Report_Template.xlsx", sample_data=None):
    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Power Report"

    # Insert Logo at A1
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 330
        img.height = 100
        img.anchor = 'A1'
        ws.add_image(img)
    else:
        print(f"Image not found: {logo_path}")

    # Title at F3
    ws["G3"] = "Daily Power Generation"
    ws["G3"].font = Font(name="Arial", size=28, bold=True)
    ws["G3"].alignment = Alignment(horizontal="left")

    # Company Info at F5
    ws["G5"] = "Company:"
    ws["G5"].font = Font(bold=True)
    ws["G5"].alignment = Alignment(horizontal="right")
    ws["H5"] = company

    # Company Info at F5
    ws["G6"] = "Power Plant:"
    ws["G6"].font = Font(bold=True)
    ws["G6"].alignment = Alignment(horizontal="right")
    ws["H6"] = powerplant

    # Leave a gap and insert table headers starting from column B
    ws.append([""] * 2)  # Spacer row
    ws.append(["", "", "Date (MM-DD-YYYY)", f"Power Generated ({metric})"])

    # Insert sample data if provided
    if sample_data:
        for row in sample_data:
            ws.append(["", "", *row])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    # Save the Excel file
    wb.save(output_file)
    print(f"Excel template saved as '{output_file}'")


def read_power_report(file_path):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return None

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    report_data = {
        "company": ws["H5"].value,
        "powerplant": ws["H6"].value,
        "records": []
    }

    start_reading = False
    for row in ws.iter_rows(values_only=True):
        # Detect header row
        if row[2] == "Date (MM-DD-YYYY)" and row[3] and "Power Generated" in row[3]:
            start_reading = True
            continue

        if start_reading:
            if row[2] and row[3]:
                report_data["records"].append({
                    "date": row[2],
                    "power_generated": row[3]
                })

    return report_data