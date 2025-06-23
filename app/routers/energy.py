from fastapi import APIRouter, Depends, Query, HTTPException, Form, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import Optional, List, Dict, Any
from decimal import Decimal
from sqlalchemy.orm import Session
from sqlalchemy import text, update, select, bindparam, ARRAY, String, Integer
from app.bronze.crud import EnergyRecords
from app.bronze.schemas import EnergyRecordOut, AddEnergyRecord
from app.public.models import RecordStatus
from app.dependencies import get_db
from app.crud.base import get_one, get_all, get_many, get_many_filtered, get_one_filtered
from datetime import datetime
import pandas as pd
import io
import math
import logging
import traceback
from collections import defaultdict
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime
from typing import Literal
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime


router = APIRouter()

def process_status_change(
    db: Session,
    energy_id: str,
    checker_id: str,
    remarks: str,
    action: str
):
    # Step 1: Validate action
    action = action.lower()
    if action not in {"approve", "revise"}:
        raise HTTPException(status_code=400, detail="Invalid action. Must be 'approve' or 'revise'.")

    # Step 2: Verify record exists
    record = db.query(EnergyRecords).filter(EnergyRecords.energy_id == energy_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Energy record not found.")

    # Step 3: Get latest status
    latest_status = (
        db.query(RecordStatus)
        .filter(RecordStatus.record_id == energy_id)
        .order_by(RecordStatus.status_timestamp.desc())
        .first()
    )
    if not latest_status:
        raise HTTPException(status_code=404, detail="Checker status not found.")
    
    current_status = latest_status.status_id

    # Step 4: Define transitions
    approve_transitions = {
        None: "URS",
        "FRS": "URS",
        "URS": "URH",
        "FRH": "URH",
        "URH": "APP",
    }

    reject_transitions = {
        "URS": "FRS",
        "URH": "FRH",
    }

    # Step 5: Determine next status
    if action == "approve":
        next_status = approve_transitions.get(current_status)
    else:  # revise
        next_status = reject_transitions.get(current_status)

    if not next_status:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot perform '{action}' from status '{current_status}'."
        )

    latest_status.status_id = next_status
    latest_status.status_timestamp = datetime.now()
    latest_status.remarks = remarks

    db.commit()
    db.refresh(latest_status)

    return {
        "message": f"Status updated to '{next_status}' via '{action}'.",
        "data": {
            "cs_id": latest_status.cs_id,
            "record_id": latest_status.record_id,
            "status_id": latest_status.status_id,
            "timestamp": latest_status.status_timestamp,
            "remarks": latest_status.remarks
        }
    }


@router.get("/energy_record", response_model=EnergyRecordOut)
def get_energy_record(
    energy_id: str = Query(..., description="ID of the energy record"),
    company: Optional[str] = Query(None, description="Filter by company"),
    powerplant: Optional[str] = Query(None, description="Filter by powerplant"),
    db: Session = Depends(get_db),
):
    filters = {"energy_id": energy_id}
    if company:
        filters["company"] = company
    if powerplant:
        filters["powerplant"] = powerplant
    
    record = get_one_filtered(db, EnergyRecords, filters)
    if not record:
        raise HTTPException(status_code=404, detail="Energy record not found")
    return record




# ====================== energy records by status ====================== #
@router.get("/energy_records_by_status", response_model=List[dict])
def get_energy_records_by_status(
    status_id: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    try:
        logging.info(f"Fetching energy records. Filter status_id: {status_id}")

        query = text("""
                SELECT     
                    er.energy_id,
                    er.power_plant_id,
                    er.date_generated::date AS date_generated,
                    er.energy_generated_kwh,
                    er.co2_avoidance_kg, 
                    pp.*, rs.status_id, st.status_name, rs.remarks
                FROM silver.csv_energy_records er
                JOIN gold.dim_powerplant_profile pp 
                    ON pp.power_plant_id = er.power_plant_id
                JOIN record_status rs on rs.record_id = er.energy_id
                JOIN public.status st on st.status_id = rs.status_id
                ORDER BY er.create_at DESC, er.date_generated DESC, er.updated_at DESC;
        """)

        result = db.execute(query, {"status_id": status_id})
        data = [dict(row._mapping) for row in result]

        logging.info(f"Returned {len(data)} records")
        return data

    except Exception as e:
        logging.error(f"Error retrieving energy records: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Internal server error")

# ====================== fact_table (func from pg) ====================== #
@router.get("/fact_energy", response_model=List[dict])
def get_fact_energy(
    power_plant_ids: Optional[List[str]] = Query(None, alias="p_power_plant_id"),
    company_ids: Optional[List[str]] = Query(None, alias="p_company_id"),
    generation_sources: Optional[List[str]] = Query(None, alias="p_generation_source"),
    provinces: Optional[List[str]] = Query(None, alias="p_province"),
    months: Optional[List[int]] = Query(None, alias="p_month"),
    quarters: Optional[List[int]] = Query(None, alias="p_quarter"),
    years: Optional[List[int]] = Query(None, alias="p_year"),
    db: Session = Depends(get_db)
):
    try:
        sql = text("""
            SELECT * FROM gold.func_fact_energy(
                p_power_plant_id := :power_plant_ids,
                p_company_id := :company_ids,
                p_generation_source := :generation_sources,
                p_province := :provinces,
                p_month := :months,
                p_quarter := :quarters,
                p_year := :years
            )
        """)

        # Convert None to NULL for Postgres array params
        params = {
            "power_plant_ids": power_plant_ids if power_plant_ids else None,
            "company_ids": company_ids if company_ids else None,
            "generation_sources": generation_sources if generation_sources else None,
            "provinces": provinces if provinces else None,
            "months": months if months else None,
            "quarters": quarters if quarters else None,
            "years": years if years else None,
        }

        result = db.execute(sql, params)
        data = [dict(row._mapping) for row in result]

        return data

    except Exception as e:
        logging.error(f"Error calling func_fact_energy: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Internal server error")


# ====================== dashboard ====================== #

@router.get("/energy_dashboard_raw", response_model=List[dict])
def get_energy_dashboard_raw(
    power_plant_ids: Optional[List[str]] = Query(None, alias="p_power_plant_id"),
    company_ids: Optional[List[str]] = Query(None, alias="p_company_id"),
    generation_sources: Optional[List[str]] = Query(None, alias="p_generation_source"),
    provinces: Optional[List[str]] = Query(None, alias="p_province"),
    months: Optional[List[int]] = Query(None, alias="p_month"),
    quarters: Optional[List[int]] = Query(None, alias="p_quarter"),
    years: Optional[List[int]] = Query(None, alias="p_year"),
    x: Optional[str] = Query(None, alias="p_x"),  # grouping by e.g., power_plant_id
    y: Optional[str] = Query(None, alias="p_y"),  # time granularity: monthly, quarterly, yearly
    v: Optional[str] = Query("energy_generated_mwh", alias="p_v"),  # aggregated metric
    db: Session = Depends(get_db)
):
    try:
        logging.info("Fetching raw energy records.")

        query = text("""
            SELECT * 
            FROM gold.func_fact_energy(                
                :power_plant_ids,
                :company_ids,
                :generation_sources,
                :provinces,
                :months,
                :quarters,
                :years
            );
        """)

        result = db.execute(query, {
            "power_plant_ids": power_plant_ids,
            "company_ids": company_ids,
            "generation_sources": generation_sources,
            "provinces": provinces,
            "months": months,
            "quarters": quarters,
            "years": years
        })

        rows = result.fetchall()
        columns = result.keys()
        data = [dict(zip(columns, row)) for row in rows]

        if not x or not y or not v:
            # Return raw data if no pivot requested
            return data

        df = pd.DataFrame(data)

        # Define time group column based on y
        if y == "monthly":
            df["time_group"] = df["year"].astype(str) + "-" + df["month"].astype(str).str.zfill(2)
        elif y == "quarterly":
            df["time_group"] = df["year"].astype(str) + "-Q" + df["quarter"].astype(str)
        elif y == "yearly":
            df["time_group"] = df["year"].astype(str)
        else:
            raise HTTPException(status_code=400, detail="Invalid y parameter. Use monthly, quarterly, or yearly.")

        if x not in df.columns or v not in df.columns:
            raise HTTPException(status_code=400, detail="Invalid x or v parameter.")

        # Perform pivot: rows = x, columns = time_group, values = v
        pivot = df.pivot_table(
            index=x,
            columns="time_group",
            values=v,
            aggfunc="sum",
            fill_value=0
        ).reset_index()

        # Convert pivot result to records
        pivot_result = pivot.to_dict(orient="records")

        logging.info(f"Pivot result: {len(pivot_result)} rows.")
        return pivot_result

    except Exception as e:
        logging.error(f"Error retrieving or processing records: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Internal server error")

    
# ====================== dashboard ====================== #
def process_query_data(
    db: Session,
    query_str: str,  
    x: str,
    y: str,
    v: List[str],
    power_plant_ids: Optional[List[str]] = None,
    company_ids: Optional[List[str]] = None,
    generation_sources: Optional[List[str]] = None,
    provinces: Optional[List[str]] = None,
    months: Optional[List[int]] = None,
    quarters: Optional[List[int]] = None,
    years: Optional[List[int]] = None
) -> Dict[str, Any]:
    # Execute query
    query = text(query_str)
    result = db.execute(query, {
        "power_plant_ids": power_plant_ids,
        "company_ids": company_ids,
        "generation_sources": generation_sources,
        "provinces": provinces,
        "months": months,
        "quarters": quarters,
        "years": years
    })

    df = pd.DataFrame(result.fetchall(), columns=result.keys())
    if df.empty:
        return {}

    if "month_name" in df.columns and pd.api.types.is_string_dtype(df["month_name"]):
        df["month_name"] = df["month_name"].str.strip()

    # Build period column
    if y == "monthly":
        df["period"] = df["year"].astype(str) + "-" + df["month"].astype(str).str.zfill(2)
    elif y == "quarterly":
        df["period"] = df["year"].astype(str) + "-Q" + df["quarter"].astype(str)
    elif y == "yearly":
        df["period"] = df["year"].astype(str)
    else:
        raise HTTPException(status_code=400, detail="Invalid y value. Use 'monthly', 'quarterly', or 'yearly'.")

    if x not in df.columns:
        raise HTTPException(status_code=400, detail=f"'{x}' not found in data columns.")
    for metric in v:
        if metric not in df.columns:
            raise HTTPException(status_code=400, detail=f"'{metric}' not found in data columns.")

    fixed_cols = {x, "period", *v}
    other_cols = [col for col in df.columns if col not in fixed_cols]

    agg_dict = {metric: 'sum' for metric in v}
    for col in other_cols:
        agg_dict[col] = 'first'

    grouped_df = df.groupby([x, "period"], dropna=False).agg(agg_dict).reset_index()

    # Line chart
    line_graph = {
        metric: [
            {
                "name": key,
                "data": [{"x": row["period"], "y": float(row[metric])} for _, row in group.iterrows()]
            }
            for key, group in grouped_df.groupby(x)
        ]
        for metric in v
    }

    # Pie chart
    pie_chart = {
        metric: [
            {
                "name": row[x],
                "value": float(row[metric]),
                "percent": float(row[metric]) / float(total) * 100 if total else 0
            }
            for _, row in grouped_df.groupby(x, dropna=False)[metric]
                .sum()
                .reset_index()
                .pipe(lambda df: (
                    df.assign(_total=df[metric].sum())  # add total column
                ))
                .iterrows()
        ]
        for metric in v
        for total in [grouped_df.groupby(x, dropna=False)[metric].sum().sum()]  # compute total once per metric
    }

    # Bar chart
    bar_chart = {
        metric: [
            {"name": row[x], "value": float(row[metric])}
            for _, row in grouped_df.groupby(x, dropna=False)[metric].sum()
            .reset_index().sort_values(metric, ascending=False).iterrows()
        ]
        for metric in v
    }

    # Step 1: Group by [x, period] and sum values in v (like energy)
    grouped_df = df.groupby([x, "period"], dropna=False)[v].sum().reset_index()

    # Step 2: Pivot to get periods as rows, x categories as columns (stacked bars)
    pivot_df = grouped_df.pivot(index="period", columns=x, values=v[0]).fillna(0).reset_index()

    # Step 3: Convert to list of dicts (for charting libraries)
    stacked_bar_chart = pivot_df.to_dict(orient="records")

    # Totals
    totals = {metric: float(df[metric].sum()) for metric in v}

    return {
        "line_graph": line_graph,
        "bar_chart": bar_chart,
        "pie_chart": pie_chart,
        "totals": totals,
        "stacked_bar":stacked_bar_chart
    }

def process_fa_data(
    db: Session,
    query_str: str,
    x: str,
    y: str,
    v: List[str],
    power_plant_ids: Optional[List[str]] = None,
    company_ids: Optional[List[str]] = None,
    ff_id: Optional[List[str]] = None,
    ff_category: Optional[List[str]] = None,
    months: Optional[List[int]] = None,
    years: Optional[List[int]] = None
) -> Dict[str, Any]:
    # Execute query
    query = text(query_str)
    result = db.execute(query, {
        "power_plant_ids": power_plant_ids,
        "company_ids": company_ids,
        "ff_id": ff_id,
        "ff_category": ff_category,
        "months": months,
        "years": years
    })

    df = pd.DataFrame(result.fetchall(), columns=result.keys())
    if df.empty:
        return {}

    if "month_name" in df.columns and pd.api.types.is_string_dtype(df["month_name"]):
        df["month_name"] = df["month_name"].str.strip()

    # Build period column
    if y == "monthly":
        df["period"] = df["year"].astype(str) + "-" + df["month"].astype(str).str.zfill(2)
    elif y == "quarterly":
        df["period"] = df["year"].astype(str) + "-Q" + df["quarter"].astype(str)
    elif y == "yearly":
        df["period"] = df["year"].astype(str)
    else:
        raise HTTPException(status_code=400, detail="Invalid y value. Use 'monthly', 'quarterly', or 'yearly'.")

    if x not in df.columns:
        raise HTTPException(status_code=400, detail=f"'{x}' not found in data columns.")
    for metric in v:
        if metric not in df.columns:
            raise HTTPException(status_code=400, detail=f"'{metric}' not found in data columns.")

    fixed_cols = {x, "period", *v}
    other_cols = [col for col in df.columns if col not in fixed_cols]

    agg_dict = {metric: 'sum' for metric in v}
    for col in other_cols:
        agg_dict[col] = 'first'

    grouped_df = df.groupby([x, "period"], dropna=False).agg(agg_dict).reset_index()

    # Line chart
    line_graph = {
        metric: [
            {
                "name": key,
                "data": [{"x": row["period"], "y": float(row[metric])} for _, row in group.iterrows()]
            }
            for key, group in grouped_df.groupby(x)
        ]
        for metric in v
    }

    # Pie chart
    pie_chart = {
        metric: [
            {
                "name": row[x],
                "value": float(row[metric]),
                "percent": float(row[metric]) / float(total) * 100 if total else 0
            }
            for _, row in grouped_df.groupby(x, dropna=False)[metric]
                .sum()
                .reset_index()
                .pipe(lambda df: (
                    df.assign(_total=df[metric].sum())  # add total column
                ))
                .iterrows()
        ]
        for metric in v
        for total in [grouped_df.groupby(x, dropna=False)[metric].sum().sum()]  # compute total once per metric
    }

    # Bar chart
    bar_chart = {
        metric: [
            {"name": row[x], "value": float(row[metric])}
            for _, row in grouped_df.groupby(x, dropna=False)[metric].sum()
            .reset_index().sort_values(metric, ascending=False).iterrows()
        ]
        for metric in v
    }

    # Totals
    totals = {metric: float(df[metric].sum()) for metric in v}

    return {
        "line_graph": line_graph,
        "bar_chart": bar_chart,
        "pie_chart": pie_chart,
        "totals": totals
    }


def to_nullable_list(param):
    return param if param else None

def process_raw_data(
    db: Session,
    query_str: str,
    power_plant_ids: Optional[List[str]] = None,
    company_ids: Optional[List[str]] = None,
    generation_sources: Optional[List[str]] = None,
    provinces: Optional[List[str]] = None,
    months: Optional[List[int]] = None,
    quarters: Optional[List[int]] = None,
    years: Optional[List[int]] = None
) -> List[Dict[str, Any]]:

    from sqlalchemy.dialects.postgresql import ARRAY

    def nullable(val):
        return val if val else None

    query = text(query_str).bindparams(
        bindparam("power_plant_ids", type_=ARRAY(String)),
        bindparam("company_ids", type_=ARRAY(String)),
        bindparam("generation_sources", type_=ARRAY(String)),
        bindparam("provinces", type_=ARRAY(String)),
        bindparam("months", type_=ARRAY(Integer)),
        bindparam("quarters", type_=ARRAY(Integer)),
        bindparam("years", type_=ARRAY(Integer)),
    )

    result = db.execute(query, {
        "power_plant_ids": nullable(power_plant_ids),
        "company_ids": nullable(company_ids),
        "generation_sources": nullable(generation_sources),
        "provinces": nullable(provinces),
        "months": nullable(months),
        "quarters": nullable(quarters),
        "years": nullable(years),
    })

    df = pd.DataFrame(result.fetchall(), columns=result.keys())
    if df.empty:
        return []

    for col in df.columns:
        if pd.api.types.is_object_dtype(df[col]):
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

    return df.to_dict(orient="records")



def parse_comma_separated(value: Optional[str]) -> Optional[List[str]]:
    if value:
        return [v.strip() for v in value.split(",") if v.strip()]
    return None

def normalize_list(lst):
    if lst is None or len(lst) == 0:
        return None
    return lst

@router.get("/energy_dashboard", response_model=Dict[str, Any])
def get_energy_dashboard(
    p_company_id: Optional[str] = Query(None),
    p_power_plant_id: Optional[str] = Query(None),
    p_generation_source: Optional[str] = Query(None),
    p_province: Optional[str] = Query(None),
    p_month: Optional[str] = Query(None),
    p_quarter: Optional[str] = Query(None),
    p_year: Optional[str] = Query(None),
    x: str = Query("company_id"),
    y: str = Query("monthly"),
    db: Session = Depends(get_db)
):
    # Parse and normalize all filter parameters
    company_ids = normalize_list(parse_comma_separated(p_company_id))
    power_plant_ids = normalize_list(parse_comma_separated(p_power_plant_id))
    generation_sources = normalize_list(parse_comma_separated(p_generation_source))
    provinces = normalize_list(parse_comma_separated(p_province))
    months = normalize_list([int(m) for m in parse_comma_separated(p_month) or []])
    quarters = normalize_list([int(q) for q in parse_comma_separated(p_quarter) or []])
    years = normalize_list([int(y) for y in parse_comma_separated(p_year) or []])

    logging.info(f"Filters - company_ids: {company_ids}, power_plant_ids: {power_plant_ids}, "
                 f"generation_sources: {generation_sources}, provinces: {provinces}, "
                 f"months: {months}, quarters: {quarters}, years: {years}")

    try:
        energy = """
            SELECT 
                *
            FROM gold.func_fact_energy(
                :power_plant_ids,
                :company_ids,
                :generation_sources,
                :provinces,
                :months,
                :quarters,
                :years
            );
        """
        energy_result = process_query_data(
            db=db,
            query_str=energy,
            x=x,
            y=y,
            v=["total_energy_generated", "total_co2_avoidance"],
            power_plant_ids=power_plant_ids,
            company_ids=company_ids,
            generation_sources=generation_sources,
            provinces=provinces,
            months=months,
            quarters=quarters,
            years=years
        )
        
        # format ------------------
        def format_large_number(value):
            if value >= 1_000_000_000:
                return f"{value / 1_000_000_000:.1f}B"
            elif value >= 1_000_000:
                return f"{value / 1_000_000:.1f}M"
            elif value <1:
                return f"{value:,.4f}"
            else:
                return math.ceil(value)

        def format_equivalence(record):
            record["co2_equivalent"] = format_large_number(round(float(record["co2_equivalent"]), 4))
            return record

        equivalence = """
            SELECT 
                energy_generated,
                co2_avoided,
                conversion_value,
                co2_equivalent,
                metric,
                equivalence_category,
                equivalence_label
            FROM gold.func_co2_equivalence_per_metric(
                :power_plant_ids,
                :company_ids,
                :generation_sources,
                :provinces,
                :months,
                :quarters,
                :years
            );
        """

        eq_result = process_raw_data(
            db=db,
            query_str=equivalence,
            power_plant_ids=power_plant_ids,
            company_ids=company_ids,
            generation_sources=generation_sources,
            provinces=provinces,
            months=months,
            quarters=quarters,
            years=years
        )

        # Nested grouping: category → EQ → list of records
        grouped_equivalence = defaultdict(lambda: defaultdict(list))

        for i, record in enumerate(eq_result):
            category = record.get("equivalence_category", "Uncategorized")  # Change this if your key is different
            eq_key = f"EQ_{i+1}"

            formatted_record = format_equivalence(record)

            grouped_equivalence[category][eq_key].append(formatted_record)

        # Convert defaultdicts to regular dicts
        equivalence_dict = {
            category: dict(eqs)
            for category, eqs in grouped_equivalence.items()
        }
        
        
        hp = """
            SELECT *
            FROM gold.func_household_powered(
                (:power_plant_ids)::VARCHAR(10)[],
                (:company_ids)::VARCHAR(10)[],
                (:generation_sources)::TEXT[],
                (:provinces)::VARCHAR(30)[],
                (:months)::INT[],
                (:quarters)::INT[],
                (:years)::INT[]
            );
        """

        hp_result = process_query_data(
            db=db,
            query_str=hp,
            x=x,
            y=y,
            v=["est_house_powered"],
            power_plant_ids=power_plant_ids,
            company_ids=company_ids,
            generation_sources=generation_sources,
            provinces=provinces,
            months=months,
            quarters=quarters,
            years=years
        )


        label_query = text("""
            SELECT
                generation_source AS label,
                TRIM(
                    CONCAT(
                        'kg CO₂ avoided = EG(kWh) × ', kg_co2_per_kwh, ' kg CO₂/kWh',
                        CASE 
                            WHEN co2_emitted_kg IS NOT NULL THEN 
                            '; kg CO₂ emitted = EG(kWh) × ' || ROUND(co2_emitted_kg / 1000.0, 6) || ' kg CO₂/kWh; Emission Reduction = kg CO₂ avoided - kg CO₂ emitted'
                            ELSE
                            ''
                        END
                    )
                ) AS formula
            FROM
                ref.ref_emission_factors
        """)

        # This returns each row as a dictionary-like object
        label_result = db.execute(label_query).mappings().all()
        label_dicts = [dict(row) for row in label_result]


        return {
            "energy_data": energy_result,
            "equivalence_data": equivalence_dict,
            "house_powered": hp_result,
            "formula":label_dicts
        }

    except Exception as e:
        logging.error(f"Error retrieving energy records: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Internal server error")


@router.get("/fund_allocation_dashboard", response_model=Dict[str, Any])
def get_fund_allocation(
    p_company_id: Optional[str] = Query(None),
    p_power_plant_id: Optional[str] = Query(None),
    p_ff_id: Optional[str] = Query(None),
    p_ff_category: Optional[str] = Query(None),
    p_month: Optional[str] = Query(None),
    p_year: Optional[str] = Query(None),
    x: str = Query("company_id"),
    y: str = Query("monthly"),
    db: Session = Depends(get_db)
):
    # Parse and normalize all filter parameters
    company_ids = normalize_list(parse_comma_separated(p_company_id))
    power_plant_ids = normalize_list(parse_comma_separated(p_power_plant_id))
    ff_id = normalize_list(parse_comma_separated(p_ff_id))
    ff_category = normalize_list(parse_comma_separated(p_ff_category))
    months = normalize_list([int(m) for m in parse_comma_separated(p_month) or []])
    years = normalize_list([int(y) for y in parse_comma_separated(p_year) or []])

    logging.info(f"Filters - company_ids: {company_ids}, power_plant_ids: {power_plant_ids}, "
                 f"ff_id: {ff_id}, ff_category: {ff_category}, "
                 f"months: {months}, years: {years}")

    try:
        energy =text( """
            SELECT 
                *
            FROM gold.func_fund_alloc(
                :power_plant_ids,
                :company_ids,
                :ff_id,
                :months,
                :years,
                :ff_category
            );
        """)
        result = db.execute(energy, {
                "power_plant_ids": power_plant_ids,
                "company_ids": company_ids,
                "ff_id": ff_id,
                "ff_category": ff_category,
                "months": months,
                "years": years
            })
        rows = result.mappings().all()
        df = pd.DataFrame(rows)
        v=['funds_allocated_peso']
        if df.empty:
            return {}

        if "month_name" in df.columns and pd.api.types.is_string_dtype(df["month_name"]):
            df["month_name"] = df["month_name"].str.strip()

        # Build period column
        if y == "monthly":
            df["period"] = df["year"].astype(str) + "-" + df["month"].astype(str).str.zfill(2)
        elif y == "quarterly":
            df["period"] = df["year"].astype(str) + "-Q" + df["quarter"].astype(str)
        elif y == "yearly":
            df["period"] = df["year"].astype(str)
        else:
            raise HTTPException(status_code=400, detail="Invalid y value. Use 'monthly', 'quarterly', or 'yearly'.")

        if x not in df.columns:
            raise HTTPException(status_code=400, detail=f"'{x}' not found in data columns.")
        for metric in v:
            if metric not in df.columns:
                raise HTTPException(status_code=400, detail=f"'{metric}' not found in data columns.")

        fixed_cols = {x, "period", *v}
        other_cols = [col for col in df.columns if col not in fixed_cols]

        # Remove any group-by columns from the aggregation dictionary to avoid duplicates
        groupby_cols = [x, "period", "ff_category", "ff_id", "ff_name"]

        # Create a copy of other_cols that excludes groupby columns
        agg_dict = {metric: 'sum' for metric in v}
        for col in other_cols:
            if col not in groupby_cols:
                agg_dict[col] = 'first' 

        # Now group by and reset index safely

        grouped_df = df.groupby(groupby_cols, dropna=False).agg(agg_dict).reset_index()
        overall = {}

        for item in v:
            value = {}
            total_per_item = 0

            for ff_cat in grouped_df["ff_category"].unique():
                chart_data = {
                    "stacked_by_period": [],
                    "stacked_by_ffid": [],
                    "pie": [],
                    "total":0
                }

                category_df = grouped_df[grouped_df["ff_category"] == ff_cat]

                category_df = category_df[pd.to_numeric(category_df[item], errors='coerce').notna()]
                category_df[item] = category_df[item].astype(float)

                # Define the group keys for your context
                group_keys = ["period", x, "ff_name"]

                # ✅ Remove duplicates based on grouping keys (not item itself)
                category_df = category_df.drop_duplicates(subset=group_keys)

                # ----- STACKED CHART BY PERIOD -----
                chart_data["stacked_by_period"] = []
                chart_data["stacked_by_ffid"] = []
                # ✅ Correct grouping for: x-axis = period, stack = ff_name
                ff_period_group = category_df.groupby(["period", "ff_name"])[item].sum().reset_index()

                # Extract unique values
                ff_names = sorted(category_df["ff_name"].unique())
                periods = sorted(category_df["period"].unique())

                # Initialize dict with period as key (since period is now x-axis)
                period_data = {period: {"period": period} for period in periods}

                # Fill the data
                for _, row in ff_period_group.iterrows():
                    period = row["period"]
                    ff_name = row["ff_name"]
                    val = row[item]
                    period_data[period][ff_name] = val

                # Fill missing ff_names with 0
                for data in period_data.values():
                    for ff in ff_names:
                        data.setdefault(ff, 0)

                chart_data["stacked_by_period"] = list(period_data.values())

                # --- Stacked by FF ID (x-axis: ff_id, stacked by x) ---
                ff_x_group = category_df.groupby(["ff_name", x])[item].sum().reset_index()

                # Get unique ff_ids and x values
                ff_ids = sorted(category_df["ff_name"].unique())
                x_values = sorted(category_df[x].unique())

                # Initialize dict with ff_id as key
                ff_data = {ff_id: {"ff_name": ff_id} for ff_id in ff_ids}

                for _, row in ff_x_group.iterrows():
                    ff_id = row["ff_name"]
                    x_val = row[x]
                    val = row[item]
                    ff_data[ff_id][x_val] = val

                # Fill missing x values with 0
                for data in ff_data.values():
                    for x_val in x_values:
                        data.setdefault(x_val, 0)

                chart_data["stacked_by_ffid"] = list(ff_data.values())




                # ----- PIE CHART -----
                pie_df = category_df.groupby("ff_name")[item].sum().reset_index()
                total_value = pie_df[item].sum()
                total_per_item += total_value  # <-- Add to overall total for this item
                chart_data["total"] = total_value  # <-- Save category total
                pie_df = category_df.groupby(x)[item].sum().reset_index()
                total_value = pie_df[item].sum()

                chart_data["pie"] = [
                    {
                        "name": row[x],
                        "value": row[item],
                        "percent": round((row[item] / total_value) * 100, 2) if total_value > 0 else 0
                    }
                    for _, row in pie_df.iterrows()
                ]

                # ----- TABLE DATA (pivot with row totals) -----
                pivot_df = category_df.pivot_table(
                    index=x,                # rows = values of 'x' (e.g., period, region)
                    columns="ff_name",      # columns = ff_name
                    values=item,            # values = current metric
                    aggfunc="sum",
                    fill_value=0
                ).reset_index()

                # Add total per row
                pivot_df["Total"] = pivot_df.drop(columns=[x]).sum(axis=1)

                # Format all numeric columns with peso sign and 2 decimal places
                for col in pivot_df.columns:
                    if col != x:  # Skip the index column
                        pivot_df[col] = pivot_df[col].apply(lambda v: f"₱{v:,.2f}")

                # Rename x column to human-readable label
                x_label_map = {
                    "power_plant_id": "Power Project",
                    "company_id": "Company",
                    "period": "Period",  # Add more mappings as needed
                    # ... any other mappings
                }
                pivot_df.rename(columns={x: x_label_map.get(x, x)}, inplace=True)

                # Convert to list of dicts for frontend rendering
                chart_data["tabledata"] = pivot_df.to_dict(orient="records")





                value[ff_cat] = chart_data  # Store chart data per ff_category
            value["total"] = total_per_item

            overall[item] = value  # ✅ Properly store data per item

        return {"data":overall}


    except Exception as e:
        logging.error(f"Error retrieving energy records: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Internal server error")
    
def serialize_row(row):
    return {
        key: float(value) if isinstance(value, Decimal) else value
        for key, value in row.items()
    }

@router.get("/overall_energy", response_model=Dict[str, Any])
def get_overall(db: Session = Depends(get_db)):
    try:
        energy = text("""
            SELECT
                CONCAT(fe.company_name, ' (', fe.company_id, ')') AS company_id,
                fe.total_energy_generated,
                fe.total_co2_avoided,
                hp.total_est_house_powered
                FROM (
                SELECT
                    company_id,company_name,
                    SUM(energy_generated_kwh) AS total_energy_generated,
                    SUM(co2_avoidance_tons) AS total_co2_avoided
                FROM gold.fact_energy_generated
                GROUP BY company_id, company_name
                ) fe
                JOIN (
                SELECT
                    company_id,
                    SUM(est_house_powered) AS total_est_house_powered
                FROM gold.func_household_powered()
                GROUP BY company_id
                ) hp
                ON fe.company_id = hp.company_id;
        """)

        result = db.execute(energy)
        rows = result.mappings().all()

        return {"data": [serialize_row(row) for row in rows]}  # ✅ Wrap in dict

    except Exception as e:
        logging.error(f"Error retrieving energy records: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Internal server error")

# ====================== template ====================== #
@router.get("/download_template", response_class=StreamingResponse)
def download_template(
    company_id: str = Query(..., description="Company ID"),
    powerplant_id: str = Query(..., description="Power Plant ID"),
    metric: Literal["kWh", "MWh", "GWh"] = Query(..., description="Metric unit (kWh, MWh, GWh)"),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    today_str = datetime.now().strftime("%Y-%m-%d")
    ws.title = f"{company_id} - {powerplant_id} - {today_str}"

    # Title at C3
    ws["C3"] = "Daily Power Generation"
    ws["C3"].font = Font(name="Arial", size=28, bold=True)
    ws["C3"].alignment = Alignment(horizontal="left")

    # Company Info
    ws["C5"] = "Company:"
    ws["C5"].font = Font(bold=True)
    ws["C5"].alignment = Alignment(horizontal="right")
    ws["D5"] = company_id

    # Power Plant Info
    ws["C6"] = "Power Plant:"
    ws["C6"].font = Font(bold=True)
    ws["C6"].alignment = Alignment(horizontal="right")
    ws["D6"] = powerplant_id

    # Metric Info
    ws["H6"] = "Metric:"
    ws["H6"].font = Font(bold=True)
    ws["H6"].alignment = Alignment(horizontal="right")
    ws["I6"] = metric

    # Headers
    ws["C8"] = "Date (MM-DD-YYYY)"
    ws["D8"] = f"Power Generated ({metric})"
    ws["C8"].font = Font(bold=True)
    ws["D8"].font = Font(bold=True)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=power_report_template.xlsx"}
    )


@router.post("/read_template", response_model=Dict[str, Any])
async def read_template(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if file is None:
        raise HTTPException(status_code=400, detail="No file uploaded.")

    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    try:
        contents = await file.read()

        if not contents:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")

        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        company = ws["D5"].value
        powerplant = ws["D6"].value
        metric = ws["I6"].value

        # Metric validation
        allowed_metrics = {"kWh", "MWh", "GWh"}
        if metric not in allowed_metrics:
            raise HTTPException(status_code=400, detail=f"Invalid metric: {metric}. Allowed: {', '.join(allowed_metrics)}")

        # Fetch existing dates from the database
        existing_dates = db.query(EnergyRecords.datetime).filter(
            EnergyRecords.power_plant_id == powerplant,
            EnergyRecords.unit_of_measurement == metric
        ).all()
        existing_dates_set = {r.datetime.date() if hasattr(r.datetime, "date") else r.datetime for r in existing_dates}

        data = []
        seen_dates = set()
        row = 9
        while True:
            date_cell = ws[f"C{row}"].value
            value_cell = ws[f"D{row}"].value

            if date_cell is None and value_cell is None:
                break

            # Validate and parse date
            try:
                if isinstance(date_cell, str):
                    parsed_date = datetime.strptime(date_cell, "%Y-%m-%d").date()
                elif isinstance(date_cell, datetime):
                    parsed_date = date_cell.date()
                elif isinstance(date_cell, (int, float)):
                    from openpyxl.utils.datetime import from_excel
                    parsed_date = from_excel(date_cell, ws.parent.epoch).date()
                else:
                    raise ValueError
            except Exception:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid or unrecognized date format in row {row}: {date_cell}"
                )

            if parsed_date in seen_dates:
                raise HTTPException(status_code=400, detail=f"Duplicate date in uploaded file at row {row}: {parsed_date}")
            seen_dates.add(parsed_date)

            if parsed_date in existing_dates_set:
                raise HTTPException(status_code=400, detail=f"Duplicate date found in database at row {row}: {parsed_date}")

            # Validate Power Generated
            if value_cell is None:
                raise HTTPException(status_code=400, detail=f"Missing 'Power Generated' value in row {row}")
            try:
                value = float(value_cell)
                if value < 0:
                    raise ValueError
            except:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid or non-numerical 'Power Generated' value in row {row}: {value_cell}"
                )

            data.append({
                "Date Generated": parsed_date.strftime("%m-%d-%Y"),
                "Power Generated": value,
                "Metric": metric
            })

            row += 1

        return {
            "company": company,
            "powerplant": powerplant,
            "columns": ["Date Generated", "Power Generated", "Metric"],
            "rows": data
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read file: {str(e)}")



@router.post("/upload_energy_file")
async def upload_energy_file(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if file is None:
        raise HTTPException(status_code=400, detail="No file uploaded.")
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    try:
        contents = await file.read()
        if not contents:
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")

        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        power_plant_id = ws["D6"].value
        unit = ws["I6"].value

        if not power_plant_id or not unit:
            raise HTTPException(status_code=400, detail="Missing required metadata in cells D6 or I6.")

        # Parse records from C9/D9
        data = []
        row = 9
        while True:
            date_cell = ws[f"C{row}"].value
            value_cell = ws[f"D{row}"].value
            if date_cell is None and value_cell is None:
                break
            if date_cell and value_cell is not None:
                data.append({
                    "date": date_cell.date() if hasattr(date_cell, "date") else date_cell,
                    "power_generated": value_cell
                })
            row += 1

        if not data:
            raise HTTPException(status_code=400, detail="No valid data rows found in the file.")

        # Prepare IDs and insert
        base_date_str = data[0]["date"].strftime("%Y%m%d")
        existing_ids = db.query(EnergyRecords.energy_id).filter(
            EnergyRecords.energy_id.like(f"EN-{base_date_str}-%")
        ).all()
        existing_suffixes = {
            int(eid[0].split("-")[-1]) for eid in existing_ids if eid[0].split("-")[-1].isdigit()
        }

        now = datetime.now()
        current_suffix = 1
        records_to_add = []
        duplicate_dates = []

        for entry in data:
            # Check if a record for this power_plant_id and date already exists
            exists = db.query(EnergyRecords).filter(
                EnergyRecords.power_plant_id == power_plant_id,
                EnergyRecords.datetime == datetime.combine(entry["date"], datetime.min.time())
            ).first()
            if exists:
                duplicate_dates.append(str(entry["date"]))
                continue  # Skip this entry

            # Find next available unique suffix
            while current_suffix in existing_suffixes:
                current_suffix += 1

            energy_id = f"EN-{base_date_str}-{current_suffix:03d}"
            current_suffix += 1  # Move to next potential suffix

            record = EnergyRecords(
                energy_id=energy_id,
                power_plant_id=power_plant_id,
                datetime=datetime.combine(entry["date"], datetime.min.time()),
                energy_generated=entry["power_generated"],
                unit_of_measurement=unit.lower(),
                create_at=now,
                updated_at=now
            )

            # Create corresponding record status
            new_log = RecordStatus(
                cs_id="CS-" + energy_id,
                record_id=energy_id,
                status_id="URS",
                status_timestamp=now,
                remarks=f"uploaded from template {file.filename} date: {now.strftime('%Y-%m-%d %H:%M:%S')}"
            )
            db.add(new_log)
            db.add(record)
            records_to_add.append(energy_id)

        if not records_to_add:
            raise HTTPException(
                status_code=400,
                detail=f"No valid records to add. Duplicate dates: {', '.join(duplicate_dates)}" if duplicate_dates else "No valid records to add."
            )

        db.commit()
        # Call stored procedure
        try:
            db.execute(text("CALL silver.load_csv_silver();"))
            db.commit()
        except Exception as proc_err:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Stored procedure error: {str(proc_err)}")

        return {
            "message": "Energy data uploaded successfully",
            "records_saved": len(records_to_add),
            "energy_ids": records_to_add,
            "duplicates_skipped": duplicate_dates
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")



# ====================== generate energy id ====================== #
def generate_energy_id(db: Session) -> str:
    today = datetime.now().strftime("%Y%m%d")
    like_pattern = f"EN-{today}-%"

    # Count how many records already exist for today
    count_today = (
        db.query(EnergyRecords)
        .filter(EnergyRecords.energy_id.like(like_pattern))
        .count()
    )

    seq = f"{count_today + 1:03d}"  # Format as 3-digit sequence
    return f"EN-{today}-{seq}"

# ====================== generate cs id ====================== #
def generate_cs_id(db: Session) -> str:
    today = datetime.now().strftime("%Y%m%d")
    like_pattern = f"CS{today}%"

    # Count how many records already exist for today
    count_today = (
        db.query(RecordStatus)
        .filter(RecordStatus.cs_id.like(like_pattern))
        .count()
    )

    seq = f"{count_today + 1:03d}"  # Format as 3-digit sequence
    return f"CS{today}{seq}"

# ====================== single add energy record ====================== #

@router.post("/add")
def add_energy_record(
    powerPlant: str = Form(...),
    date: str = Form(...),
    energyGenerated: float = Form(...),
    checker: str = Form(...),
    metric: str = Form(...),
    remarks: str = Form(...),
    db: Session = Depends(get_db),
):
    
    try:
        # Parse date
        try:
            parsed_date = datetime.strptime(date, "%Y-%m-%d")
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=f"Date parsing error: {str(ve)}")

        # Check for existing record
        try:
            existing = db.query(EnergyRecords).filter(
                EnergyRecords.power_plant_id == powerPlant,
                EnergyRecords.datetime == parsed_date
            ).first()
            if existing:
                raise HTTPException(
                    status_code=400,
                    detail={
                        "type": "duplicate_error",
                        "message": f"Duplicate record: energy data for {powerPlant} on {parsed_date.strftime('%Y-%m-%d')} already exists."
                    }
                )
        except Exception as db_err:
            raise HTTPException(status_code=500, detail=f"Database query failed: {str(db_err)}")

        # Generate ID
        try:
            new_id = generate_energy_id(db)
        except Exception as id_err:
            raise HTTPException(status_code=500, detail=f"ID generation failed: {str(id_err)}")

        # Create record and log
        try:
            new_record = EnergyRecords(
                energy_id=new_id,
                power_plant_id=powerPlant,
                datetime=parsed_date,
                energy_generated=energyGenerated,
                unit_of_measurement=metric,
            )
            new_log = RecordStatus(
                cs_id="CS-" + new_id,
                record_id=new_id,
                status_id="URS",
                status_timestamp=datetime.now(),
                remarks=remarks
            )

            db.add(new_record)
            db.add(new_log)
            db.commit()
            db.refresh(new_record)
        except Exception as record_err:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Failed to insert record or log: {str(record_err)}")

        # Call stored procedure
        try:
            db.execute(text("CALL silver.load_csv_silver();"))
            db.commit()
        except Exception as proc_err:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Stored procedure error: {str(proc_err)}")

        return {
            "message": "Energy record successfully added.",
            "data": {
                "energy_id": new_record.energy_id,
                "power_plant_id": new_record.power_plant_id,
                "datetime": new_record.datetime,
                "energy_generated": new_record.energy_generated,
                "unit_of_measurement": new_record.unit_of_measurement
            }
        }

    except HTTPException as http_err:
        raise http_err

    except Exception as e:
        db.rollback()
        tb = traceback.format_exc()
        raise HTTPException(
            status_code=500,
            detail=f"Unexpected error: {str(e)}\nTraceback:\n{tb}"
        )



# ====================== bulk add energy record ====================== #
@router.post("/bulk_add")
def bulk_add_energy_record(
    # powerPlant: str = Form(...),
    # checker: str = Form(...),
    # metric: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if file.content_type not in [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel"
    ]:
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload an Excel file.")

    try:
        contents = file.file.read()
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")

    if 'date' not in df.columns or 'energy_generated' not in df.columns or 'powerPlant' not in df.columns or 'metric' not in df.columns:
        raise HTTPException(status_code=400, detail="Excel must contain 'date', 'energy_generated', 'powerPlant', and 'metric' columns.")

    # Generate base energy_id prefix for bronze.energy records
    first_id = generate_energy_id(db)
    parts = first_id.split("-")
    prefix = "-".join(parts[:2])
    counter = int(parts[-1])
    
    # Generate base cs_id prefix for status
    first_cs_id = generate_cs_id(db)
    cs_prefix = first_cs_id[:-3]  # except the last 3 digits
    cs_counter = int(first_cs_id[-3:])  # last 3 digits

    inserted = 0
    updated = 0

    for i, (_, row) in enumerate(df.iterrows()):
        try:
            date_val = pd.to_datetime(row["date"], format="%Y-%m-%d").date()
            energy_val = float(row["energy_generated"])
            powerPlant = str(row["powerPlant"])
            metric = str(row["metric"])
        except Exception:
            raise HTTPException(status_code=400, detail=f"Invalid format at row {i+2}.")

        # Check if a record exists
        existing_query = select(EnergyRecords.energy_generated).where(
            EnergyRecords.power_plant_id == powerPlant,
            EnergyRecords.datetime == date_val
        )
        result = db.execute(existing_query).first()

        if result:
            # aggregate
            update_stmt = (
                update(EnergyRecords)
                .where(
                    EnergyRecords.power_plant_id == powerPlant,
                    EnergyRecords.datetime == date_val
                )
                .values(energy_generated=EnergyRecords.energy_generated + energy_val)
            )
            db.execute(update_stmt)
            updated += 1
        else:
            # new record
            new_id = f"{prefix}-{str(counter).zfill(3)}"
            counter += 1

            new_record = EnergyRecords(
                energy_id=new_id,
                power_plant_id=powerPlant,
                datetime=date_val,
                energy_generated=energy_val,
                unit_of_measurement=metric,
            )
            db.add(new_record)

            # Add corresponding status log
            new_log_id = f"{cs_prefix}{str(cs_counter).zfill(3)}"
            cs_counter += 1
            
            new_log = RecordStatus(
                cs_id=new_log_id,
                #checker_id=checker,
                record_id=new_id,
                status_id="URS",
                status_timestamp=datetime.now(),
                remarks="Newly Added"
            )
            db.add(new_log)

            inserted += 1

    try:
        db.commit()
        db.execute(text("CALL silver.load_csv_silver();"))
        db.commit()
        return {
            "message": "Processed successfully.",
            "inserted": inserted,
            "updated": updated
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {e}")

# ====================== update status ====================== #
@router.post("/update_status")
def change_status(
    energy_id: str = Form(...),
    checker_id: str = Form(...),
    remarks: str = Form(...),
    action: str = Form(...),
    db: Session = Depends(get_db),
):
    try:
        return process_status_change(
            db=db,
            energy_id=energy_id,
            checker_id=checker_id,
            remarks=remarks,
            action=action
        )
    except HTTPException as http_err:
        raise http_err
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

# ====================== edit energy record ====================== #

@router.post("/edit")
def edit_energy_record(
    energy_id: str = Form(...),
    powerPlant: str = Form(...),
    date: str = Form(...),
    energyGenerated: float = Form(...),
    checker: str = Form(...),
    metric: str = Form(...),
    remarks:str=Form(...),
    db: Session = Depends(get_db),
):
    try:
        parsed_date = datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")
    
    
    # Step 3: Get latest status
    latest_status = (
        db.query(RecordStatus)
        .filter(RecordStatus.record_id == energy_id)
        .order_by(RecordStatus.status_timestamp.desc())
        .first()
    )
    current_status = latest_status.status_id if latest_status else None
    new_status = "URH" if current_status in ['FRH', 'URH'] else "URS"


    # update
    update_stmt = (
        update(EnergyRecords)
        .where(EnergyRecords.energy_id == energy_id)
        .values(
            energy_generated = energyGenerated,
            unit_of_measurement = metric,
            updated_at = parsed_date,
            power_plant_id = powerPlant
        )
    )

    try:
        db.execute(update_stmt)

        # Step 4: Update the existing RecordStatus
        latest_status.status_id = new_status
        latest_status.status_timestamp = datetime.now()
        latest_status.remarks = remarks

        db.commit()
        return {"message": "Energy record updated successfully."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")


