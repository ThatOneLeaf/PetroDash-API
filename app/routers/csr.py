from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import List, Dict, Optional
from decimal import Decimal
import logging
import traceback
from datetime import datetime
import pandas as pd
from io import BytesIO
from app.bronze.crud import insert_csr_activity, update_csr_activity, bulk_upload_csr_activity
from fastapi.responses import StreamingResponse
import openpyxl
import io
import math

from ..dependencies import get_db

router = APIRouter()

def create_excel_template(headers: List[str], filename: str) -> BytesIO:
    """Create minimal Excel template with just headers and readable column widths"""
    df = pd.DataFrame({header: [] for header in headers})
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.sheets['Sheet1']
        
        # Auto-adjust column widths to make headers readable
        for column in worksheet.columns:
            max_length = len(str(column[0].value)) + 2  # Header length + padding
            worksheet.column_dimensions[column[0].column_letter].width = max_length
    
    output.seek(0)
    return output

@router.get("/programs", response_model=List[Dict])
def get_csr_programs(db: Session = Depends(get_db)):
    """
    Get all CSR programs
    Returns list of programs with their details
    """
    try:
        logging.info("Executing CSR programs query")
        
        result = db.execute(text("""
            SELECT 
                program_id,
                program_name,
                date_created,
                date_updated
            FROM silver.csr_programs
            WHERE (
                program_id = 'HE'
                OR program_id = 'ED'
                OR program_id = 'LI'
            )
            ORDER BY program_name
        """))
        
        data = [
            {
                'programId': row.program_id,
                'programName': row.program_name,
                'dateCreated': row.date_created.isoformat() if row.date_created else None,
                'dateUpdated': row.date_updated.isoformat() if row.date_updated else None
            }
            for row in result
        ]
        
        logging.info(f"Query returned {len(data)} CSR programs")
        return data

    except Exception as e:
        logging.error(f"Error fetching CSR programs: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/projects", response_model=List[Dict])
def get_csr_projects(program_id: Optional[str] = None, db: Session = Depends(get_db)):
    """
    Get CSR projects, optionally filtered by program_id
    Returns list of projects with their program information
    """
    try:
        logging.info(f"Executing CSR projects query with program_id filter: {program_id}")
        
        where_clause = ""
        params = {}
        
        if program_id:
            where_clause = "WHERE cp.program_id = :program_id"
            params['program_id'] = program_id
        
        result = db.execute(text(f"""
            SELECT 
                cp.program_id,
                pr.program_name,
                cp.project_id,
                cp.project_name,
                cp.project_metrics,
                cp.date_created,
                cp.date_updated
            FROM silver.csr_projects cp
            JOIN silver.csr_programs pr ON cp.program_id = pr.program_id
            {where_clause} AND (
                cp.project_id LIKE 'HE%'
                OR cp.project_id LIKE 'ED%'
                OR cp.project_id LIKE 'LI%'
            )
            ORDER BY pr.program_name, cp.project_name
        """), params)

        data = [
            {
                'projectId': row.project_id,
                'programId': row.program_id,
                'programName': row.program_name,
                'projectName': row.project_name,
                'projectMetrics': row.project_metrics,
                'dateCreated': row.date_created.isoformat() if row.date_created else None,
                'dateUpdated': row.date_updated.isoformat() if row.date_updated else None
            }
            for row in result
        ]
        
        logging.info(f"Query returned {len(data)} CSR projects")
        return data
        
    except Exception as e:
        logging.error(f"Error fetching CSR projects: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/activities", response_model=List[Dict])
def get_csr_activities(
    year: Optional[int] = None,
    company_id: Optional[str] = None,
    program_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get CSR activities with optional filters
    Returns list of activities with company, project, and program information
    """
    try:
        logging.info(f"Executing CSR activities query with filters - year: {year}, company_id: {company_id}, program_id: {program_id}")
        
        where_conditions = []
        params = {}
        
        if year:
            where_conditions.append("ca.project_year = :year")
            params['year'] = year
            
        if company_id:
            where_conditions.append("ca.company_id = :company_id")
            params['company_id'] = company_id
            
        if program_id:
            where_conditions.append("cp.program_id = :program_id")
            params['program_id'] = program_id
        
        where_clause = ""
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions)

        result = db.execute(text(f"""
            SELECT 
                ca.csr_id,
                ca.company_id,
                cm.company_name,
                ca.project_id,
                cp.project_name,
                cp.program_id,
                pr.program_name,
                ca.project_year,
                ROUND(ca.csr_report::numeric, 2) as csr_report,
                ROUND(ca.project_expenses::numeric, 2) as project_expenses,
                rs.status_id,
                ca.project_remarks,
                ca.date_created,
                ca.date_updated
            FROM silver.csr_activity ca
            JOIN ref.company_main cm ON ca.company_id = cm.company_id
            JOIN silver.csr_projects cp ON ca.project_id = cp.project_id
            JOIN silver.csr_programs pr ON cp.program_id = pr.program_id
            JOIN public.record_status rs ON ca.csr_id = rs.record_id
            {where_clause} AND
                (
                    ca.project_id LIKE 'HE%' 
                    OR ca.project_id LIKE 'ED%' 
                    OR ca.project_id LIKE 'LI%'
                )
            ORDER BY rs.status_id DESC
        """), params)

        data = [
            {
                'csrId': row.csr_id,
                'companyId': row.company_id,
                'companyName': row.company_name,
                'programId': row.program_id,
                'programName': row.program_name,
                'projectId': row.project_id,
                'projectName': row.project_name,
                'projectYear': row.project_year,
                'csrReport': float(row.csr_report) if row.csr_report else 0,
                'projectExpenses': float(row.project_expenses) if row.project_expenses else 0,
                'projectRemarks': row.project_remarks,
                'statusId': (
                    "Approved" if row.status_id == "APP"
                    else "For Revision (Site)" if row.status_id == "FRS"
                    else "For Revision (Head)" if row.status_id == "FRH"
                    else "Under Review (Site)" if row.status_id == "URS"
                    else "Under Review (Head)" if row.status_id == "URH"
                    else row.status_id
                ),
            }
            for row in result
        ]

        logging.info(f"Query returned {len(data)} CSR activities")
        return data
        
    except Exception as e:
        logging.error(f"Error fetching CSR activities: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/activities-specific", response_model=Dict)
def get_csr_activity_specific(
    csr_id: str = Query(..., alias="csrId"),
    db: Session = Depends(get_db)
):
    """
    Get a single CSR activity by project_id
    """
    try:
        result = db.execute(text("""
            SELECT 
                ca.csr_id,
                ca.company_id,
                cm.company_name,
                ca.project_id,
                cp.project_name,
                cp.program_id,
                pr.program_name,
                ca.project_year,
                ROUND(ca.csr_report::numeric, 2) as csr_report,
                ROUND(ca.project_expenses::numeric, 2) as project_expenses,
                csl.status_id,
                ca.project_remarks,
                ca.date_created,
                ca.date_updated
            FROM silver.csr_activity ca
            JOIN ref.company_main cm ON ca.company_id = cm.company_id
            JOIN silver.csr_projects cp ON ca.project_id = cp.project_id
            JOIN silver.csr_programs pr ON cp.program_id = pr.program_id
            JOIN public.record_status csl ON ca.csr_id = csl.record_id
            WHERE ca.csr_id = :csr_id
            LIMIT 1
        """), {"csr_id": csr_id})

        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="CSR activity not found")

        data = {
            'csrId': row.csr_id,
            'companyId': row.company_id,
            'companyName': row.company_name,
            'programId': row.program_id,
            'programName': row.program_name,
            'projectId': row.project_id,
            'projectName': row.project_name,
            'projectYear': row.project_year,
            'csrReport': float(row.csr_report) if row.csr_report else 0,
            'projectExpenses': float(row.project_expenses) if row.project_expenses else 0,
            'projectRemarks': row.project_remarks,
            'statusId': (
                "Approved" if row.status_id == "APP"
                else "For Revision (Site)" if row.status_id == "FRS"
                else "For Revision (Head)" if row.status_id == "FRH"
                else "Under Review (Site)" if row.status_id == "URS"
                else "Under Review (Head Level)" if row.status_id == "URH"
                else row.status_id
            ),
        }
        return data

    except Exception as e:
        logging.error(f"Error fetching CSR activity: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/help-report", response_model=List[Dict])
def get_csr_activities(
    year: Optional[int] = None,
    company_id: Optional[str] = None,
    program_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get CSR activities with optional filters
    Returns list of activities with company, project, and program information
    """
    try:
        logging.info(f"Executing CSR activities query with filters - year: {year}, company_id: {company_id}")
        
        where_conditions = []
        params = {}
        
        if year:
            where_conditions.append("ca.project_year = :year")
            params['year'] = year
            
        if company_id:
            where_conditions.append("ca.company_id = :company_id")
            params['company_id'] = company_id
        
        where_clause = ""
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions) + " AND (ca.project_id LIKE 'HE%' OR ca.project_id LIKE 'ED%' OR ca.project_id LIKE 'LI%')"
        else:
            where_clause = "WHERE (ca.project_id LIKE 'HE%' OR ca.project_id LIKE 'ED%' OR ca.project_id LIKE 'LI%') AND rs.status_id = 'APP'"

        result = db.execute(text(f"""
            SELECT 
                ca.csr_id,
                ca.company_id,
                cm.company_name,
                ca.project_id,
                cp.project_name,
                cp.program_id,
                pr.program_name,
                ca.project_year,
                ROUND(ca.csr_report::numeric, 2) as csr_report,
                ROUND(ca.project_expenses::numeric, 2) as project_expenses
            FROM silver.csr_activity ca
            JOIN ref.company_main cm ON ca.company_id = cm.company_id
            JOIN silver.csr_projects cp ON ca.project_id = cp.project_id
            JOIN silver.csr_programs pr ON cp.program_id = pr.program_id
            JOIN public.record_status rs ON ca.csr_id = rs.record_id
            {where_clause}
        """), params)

        data = [
            {
                'csrId': row.csr_id,
                'companyId': row.company_id,
                'companyName': row.company_name,
                'programId': row.program_id,
                'programName': row.program_name,
                'projectId': row.project_id,
                'projectName': row.project_name,
                'projectYear': row.project_year,
                'csrReport': float(row.csr_report) if row.csr_report else 0,
                'projectExpenses': float(row.project_expenses) if row.project_expenses else 0,
            }
            for row in result
        ]

        logging.info(f"Query returned {len(data)} CSR activities")
        return data
        
    except Exception as e:
        logging.error(f"Error fetching CSR activities: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/investments-per-project", response_model=List[Dict])
def get_help_investments(
    year: Optional[int] = None,
    company_id: Optional[str] = None,
    program_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get HELP investments with optional filters
    Returns list of expenses per company, project, and program
    """
    try:
        logging.info(f"Executing CSR investments query with filters - year: {year}, company_id: {company_id}, program_id: {program_id}")
        
        where_conditions = []
        params = {}
        
        if year:
            where_conditions.append("cact.project_year = :year")
            params['year'] = year
            
        if company_id:
            where_conditions.append("cact.company_id = :company_id")
            params['company_id'] = company_id
            
        # if program_id:
        #     where_conditions.append("cp.program_id = :program_id")
        #     params['program_id'] = program_id
        
        where_clause = ""
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions) + " AND (cact.project_id LIKE 'HE%' OR cact.project_id LIKE 'ED%' OR cact.project_id LIKE 'LI%')"
        else:
            where_clause = "WHERE (cact.project_id LIKE 'HE%' OR cact.project_id LIKE 'ED%' OR cact.project_id LIKE 'LI%') AND csl.status_id = 'APP'"

        result = db.execute(text(f"""
            SELECT 
                cact.project_id,
                cproj.project_name,
                SUM(cact.project_expenses) AS "project_investments"
            FROM silver.csr_activity AS cact
            LEFT JOIN silver.csr_projects AS cproj
            ON cact.project_id = cproj.project_id
            LEFT JOIN public.record_status csl 
            ON cact.csr_id = csl.record_id
            LEFT JOIN ref.company_main AS comp
            ON cact.company_id = comp.company_id
            {where_clause}
            GROUP BY 
                cact.project_id,
                cproj.project_name
            ORDER BY 
                project_investments DESC;
        """), params)

        data = [
            {
                'projectName': row.project_name,
                'projectExpenses': float(row.project_investments) if row.project_investments else 0
            }
            for row in result
        ]

        logging.info(f"Query returned {len(data)} CSR activities")
        return data
        
    except Exception as e:
        logging.error(f"Error fetching CSR activities: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/investments-per-program", response_model=List[Dict])
def get_help_investments(
    year: Optional[int] = None,
    company_id: Optional[str] = None,
    program_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get HELP investments with optional filters
    Returns list of expenses per company, project, and program
    """
    try:
        logging.info(f"Executing CSR investments query with filters - year: {year}, company_id: {company_id}, program_id: {program_id}")
        
        where_conditions = []
        params = {}
        
        if year:
            where_conditions.append("cact.project_year = :year")
            params['year'] = year
            
        if company_id:
            where_conditions.append("cact.company_id = :company_id")
            params['company_id'] = company_id
            
        # if program_id:
        #     where_conditions.append("cp.program_id = :program_id")
        #     params['program_id'] = program_id
        
        where_clause = ""
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions) + " AND (cact.project_id LIKE 'HE%' OR cact.project_id LIKE 'ED%' OR cact.project_id LIKE 'LI%')"
        else:
            where_clause = "WHERE cact.project_id LIKE 'HE%' OR cact.project_id LIKE 'ED%' OR cact.project_id LIKE 'LI%' AND csl.status_id = 'APP'"

        result = db.execute(text(f"""
            SELECT 
                cprog.program_name,
                SUM(project_expenses) AS "project_investments",
                cact.date_updated
            FROM silver.csr_activity AS cact
            LEFT JOIN silver.csr_projects AS cproj
            ON cact.project_id = cproj.project_id
            LEFT JOIN silver.csr_programs AS cprog
            ON cproj.program_id = cprog.program_id
            LEFT JOIN public.record_status csl 
            ON cact.csr_id = csl.record_id
            LEFT JOIN ref.company_main AS ccomp
            ON cact.company_id = ccomp.company_id
            {where_clause}
            GROUP BY cprog.program_name, cact.date_updated
            ORDER BY "project_investments"
        """), params)

        data = [
            {
                'programName': row.program_name,
                'projectExpenses': float(row.project_investments) if row.project_investments else 0,
                'dateUpdated': row.date_updated
            }
            for row in result
        ]

        logging.info(f"Query returned {len(data)} CSR activities")
        return data
        
    except Exception as e:
        logging.error(f"Error fetching CSR activities: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/investments-per-company", response_model=List[Dict])
def get_help_investments(
    year: Optional[int] = None,
    company_id: Optional[str] = None,
    program_id: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    Get HELP investments with optional filters
    Returns list of expenses per company, project, and program
    """
    try:
        logging.info(f"Executing CSR investments query with filters - year: {year}, company_id: {company_id}, program_id: {program_id}")
        
        where_conditions = []
        params = {}
        
        if year:
            where_conditions.append("cact.project_year = :year")
            params['year'] = year
            
        # if company_id:
        #     where_conditions.append("cact.company_id = :company_id")
        #     params['company_id'] = company_id
            
        # if program_id:
        #     where_conditions.append("cp.program_id = :program_id")
        #     params['program_id'] = program_id
        
        where_clause = ""
        if where_conditions:
            where_clause = "WHERE " + " AND ".join(where_conditions) + " AND (cact.project_id LIKE 'HE%' OR cact.project_id LIKE 'ED%' OR cact.project_id LIKE 'LI%')"
        else:
            where_clause = "WHERE cact.project_id LIKE 'HE%' OR cact.project_id LIKE 'ED%' OR cact.project_id LIKE 'LI%' AND csl.status_id = 'APP'"

        result = db.execute(text(f"""
            SELECT 
                ccomp.company_id,
                SUM(project_expenses) AS "project_investments"
            FROM silver.csr_activity AS cact
            LEFT JOIN silver.csr_projects AS cproj
            ON cact.project_id = cproj.project_id
            LEFT JOIN silver.csr_programs AS cprog
            ON cproj.program_id = cprog.program_id
            LEFT JOIN public.record_status csl 
            ON cact.csr_id = csl.record_id
            LEFT JOIN ref.company_main AS ccomp
            ON cact.company_id = ccomp.company_id
            {where_clause}
            GROUP BY ccomp.company_id
            ORDER BY "project_investments" DESC
        """), params)

        data = [
            {
                'companyId': row.company_id,
                'projectExpenses': float(row.project_investments) if row.project_investments else 0
            }
            for row in result
        ]

        logging.info(f"Query returned {len(data)} CSR activities")
        return data
        
    except Exception as e:
        logging.error(f"Error fetching CSR activities: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

# ----------------------- POST METHODS ----------------------------

@router.post("/activities-update")
def update_csr_activity_single(data: dict, db: Session = Depends(get_db)):
    try:
        logging.info("Update single csr activity record")
        CURRENT_YEAR = datetime.now().year

        required_fields = ['company_id', 'project_id', 'project_year', 'csr_report', 'project_expenses']
        missing = [field for field in required_fields if field not in data]
        if missing:
            # Return a consistent error response
            return {"success": False, "message": f"Missing required fields: {missing}"}

        if not isinstance(data["company_id"], str) or not data["company_id"].strip():
            return {"success": False, "message": "Invalid Company ID"}

        if not isinstance(data["project_id"], str) or not data["project_id"].strip():
            return {"success": False, "message": "Invalid project ID"}

        if not isinstance(data["project_year"], int) or (int(data["project_year"]) > CURRENT_YEAR or int(data["project_year"]) < 2000):
            return {"success": False, "message": "Invalid project year"}
        
        if not isinstance(data["csr_report"], int) or (int(data["csr_report"]) <= 0):
            return {"success": False, "message": "Invalid beneficiaries"}

        if not isinstance(data["project_expenses"], (int, float)):
            return {"success": False, "message": "Invalid project investment"}

        record = {
            "csr_id": data["csr_id"],
            "company_id": data["company_id"],
            "project_id": data["project_id"],
            "project_year": data["project_year"],
            "csr_report": data["csr_report"],
            "project_expenses": data["project_expenses"],
            "project_remarks": data["project_remarks"],
        }

        update_csr_activity(db, record)

        return {"success": True, "message": "1 record successfully inserted."}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return {"success": False, "message": str(e)}

@router.post("/activities-single")
def insert_csr_activity_single(data: dict, db: Session = Depends(get_db)):
    try:
        logging.info("Add single csr activity record")
        CURRENT_YEAR = datetime.now().year

        required_fields = ['company_id', 'project_id', 'project_year', 'csr_report', 'project_expenses']
        missing = [field for field in required_fields if field not in data]
        if missing:
            # Return a consistent error response
            return {"success": False, "message": f"Missing required fields: {missing}"}

        if not isinstance(data["company_id"], str) or not data["company_id"].strip():
            return {"success": False, "message": "Invalid company_id"}

        if not isinstance(data["project_id"], str) or not data["project_id"].strip():
            return {"success": False, "message": "Invalid project ID"}
        
        if not isinstance(data["project_year"], int) or (int(data["project_year"]) > CURRENT_YEAR or int(data["project_year"]) < 2000):
            return {"success": False, "message": "Invalid project year"}
        
        if not isinstance(data["csr_report"], int) or (int(data["csr_report"]) <= 0):
            return {"success": False, "message": "Invalid beneficiaries"}

        if not isinstance(data["project_expenses"], (int, float)) or data["project_expenses"] < 0 or pd.isna(data["project_expenses"]):
            return {"success": False, "message": "Invalid project investment"}

        record = {
            "company_id": data["company_id"],
            "project_id": data["project_id"],
            "project_year": data["project_year"],
            "csr_report": data["csr_report"],
            "project_expenses": data["project_expenses"],
            "project_remarks": data["project_remarks"]
        }
        insert_csr_activity(db, record)

        return {"success": True, "message": "1 record successfully inserted."}

    except Exception as e:
        db.rollback()
        return {"success": False, "message": str(e)}

@router.get("/help-activity-template")
async def download_help_activity_template():
    """Generate Excel template for HELP activities data"""
    headers = [
        ("company_id", "Registered Company IDs (PERC, PGEC, PSC, MGI, PWEI, ESEC, RGEC, BEP_NL, BEP_NM, BEP_EP, BGEC, SJGEC, DGEC, BKS)"),
        ("project_id", "Registered Project IDs: HE_AMM - Annual Medical Mission, HE_CHC - Community Health Center, HE_NP - Nutrition Program, HE_SA - Service Ambulance, HE_MC - Mobile Clinics, ED_AS - Adopted School, ED_EMD - Educational Mobile Devices, ED_SP - Scholarship Program, ED_TT - Teacher Training, LI_LT_T - Livelihood Training"),
        ("project_year", "Year of the project"),
        ("csr_report", "Number of beneficiaries"),
        ("project_expenses", "Amount invested for the project"),
        ("project_remarks", "For project tracking or identity (i.e: project's title, target beneficiary)")
    ]

    wb = openpyxl.Workbook()
    sheet1 = wb.active
    sheet1.title = "Sheet1"
    sheet2 = wb.create_sheet(title="project_details")

    for col, (header, _) in enumerate(headers, start=1):
        sheet1.cell(row=1, column=col, value=header)

    sheet2.cell(row=1, column=1, value="Header")
    sheet2.cell(row=1, column=2, value="Input Description")
    for row, (header, desc) in enumerate(headers, start=2):
        sheet2.cell(row=row, column=1, value=header)
        sheet2.cell(row=row, column=2, value=desc)

    # Write to in-memory buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = 'help_activity_template.xlsx'

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@router.post("/help-activity-bulk")
def bulk_help_activity(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        # raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")
        return {"success": False, "message": f"Invalid file format. Please upload an Excel file."}
    
    try:
        logging.info(f"Add bulk data")
        contents = file.file.read()
        df = pd.read_excel(BytesIO(contents))
        company_list = ['PERC', 'PGEC', 'PSC', 'MGI', 'PWEI', 'ESEC', 'RGEC', 'BEP_NL', 'BEP_NM', 'BEP_EP', 'BGEC', 'SJGEC', 'DGEC', 'BKS']
        project_ids = ['HE_AMM', 'HE_CHC', 'HE_NP', 'HE_SA', 'HE_MC', 'ED_SA', 'ED_EMD', 'ED_SP', 'ED_TT', 'LI_LT_T']

        required_columns = {'company_id', 'project_id', 'project_year', 'csr_report', 'project_expenses'}
        if not required_columns.issubset(df.columns):
            return {"success": False, "message": f"Missing required fields: {required_columns - set(df.columns)}"}

        rows = []
        validation_errors = []
        CURRENT_YEAR = datetime.now().year

        for i, row in df.iterrows():
            row_number = i + 1
            
            if not isinstance(row["company_id"], str) or not row["company_id"].strip() or row["company_id"] not in company_list:
                validation_errors.append(f"Row {row_number}: Invalid company ID")
                continue
                
            if not isinstance(row["project_id"], str) or not row["project_id"].strip() or row["project_id"] not in project_ids:
                validation_errors.append(f"Row {row_number}: Invalid project ID")

            if not isinstance(row["project_year"], (int)) or (int(row["project_year"]) > CURRENT_YEAR or int(row["project_year"]) < 2000):
                validation_errors.append(f"Row {row_number}: Invalid project year")
                continue
            
            if not isinstance(row["csr_report"], (int)) or row["csr_report"] < 0:
                validation_errors.append(f"Row {row_number}: Invalid CSR beneficiary")
                continue
            
            # Check if project_expenses is empty/NaN/invalid
            if not isinstance(row["project_expenses"], (int, float)) or row["project_expenses"] < 0 or pd.isna(row["project_expenses"]):
                validation_errors.append(f"Row {row_number}: Invalid project investments")
                continue

            # if not row["project_remarks"].strip():
            #     validation_errors.append(f"Row {row_number}: Empty project remarks")
            #     continue

            # If all validations pass, add to rows
            rows.append({
                "company_id": row["company_id"].strip(),
                "project_id": row["project_id"],
                "project_year": int(row["project_year"]),
                "csr_report": int(row["csr_report"]),
                "project_expenses": float(row["project_expenses"]),
                "project_remarks": row["project_remarks"],
            })

        # If there are validation errors, return them
        if validation_errors:
            error_message = "Data validation failed:\n" + "\n".join(validation_errors)
            raise HTTPException(status_code=422, detail=error_message)
            # return {"message": f"Missing required fields: {error_message}"}

        # If no validation errors, proceed with bulk insert
        if not rows:
            raise HTTPException(status_code=400, detail="No valid data rows found to insert")
            # return {"message": f"No valid data rows found to insert"}

        count = bulk_upload_csr_activity(db, rows)
        return {"message": f"{count} records successfully inserted."}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))